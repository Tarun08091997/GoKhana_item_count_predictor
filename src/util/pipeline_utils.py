"""
Utility functions for the pipeline: file locator, logging, Excel operations, etc.
"""

import json
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
from pymongo import MongoClient
from bson import ObjectId
# ConfigManger is now available through connection_manager for backward compatibility

# Import from new utility modules
from src.util.progress_bar import ProgressBar
from src.util.path_utils import (
    get_output_base_dir, get_input_base_dir, get_fr_data_path, get_retrain_path,
    get_pipeline_type, set_pipeline_type, get_pipeline_log_dir,
    sanitize_name, get_file_name, extract_item_name_from_filename,
    get_item_name_from_file, get_item_name_from_excel,
    get_model_file_name, get_result_file_name
)

LOGGER = logging.getLogger(__name__)

# Cache for MongoDB names
_name_cache: Dict[str, Optional[str]] = {}


def get_mongo_names(foodcourt_id: str, restaurant_id: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Fetch food court name and restaurant name from MongoDB.
    Uses ConnectionManager for centralized connection handling.
    Returns (foodcourt_name, restaurant_name) or (None, None) if not found.
    """
    try:
        from src.util.connection_manager import get_connection_manager
        conn_mgr = get_connection_manager()
        foodcourt_name, restaurant_name = conn_mgr.get_mongo_names(foodcourt_id, restaurant_id)
        return foodcourt_name, restaurant_name
    except Exception as e:
        LOGGER.warning(f"Failed to get MongoDB names: {e}")
        return None, None


def get_item_name_from_filename(file_name: str, foodcourt_id: str, restaurant_id: str, item_id: str) -> Optional[str]:
    """
    Extract item name from file name.
    Format: {foodcourt_id}_{restaurant_id}_{item_id}_{item_name}_...
    Example: 5f338dce8f277f4c2f4ac99f_65fc218c0858ce0012f959b7_65fd1da416890e00112acc96_Butter_Cookie_enrich_data.csv
    
    Args:
        file_name: File name containing the item name
        foodcourt_id: Foodcourt ID
        restaurant_id: Restaurant ID
        item_id: Item ID
    
    Returns:
        Item name if found, None otherwise
    """
    if not file_name:
        return None
    
    import re
    
    # Remove file extension
    name_without_ext = file_name.replace('.csv', '').replace('.xlsx', '')
    
    # Pattern: foodcourt_id_restaurant_id_item_id_item_name_...
    pattern = f"{foodcourt_id}_{restaurant_id}_{item_id}_(.+?)_"
    match = re.search(pattern, name_without_ext)
    
    if match:
        item_name = match.group(1)
        # Replace underscores with spaces for readability
        return item_name.replace('_', ' ')
    
    # Try alternative pattern without trailing underscore
    pattern2 = f"{foodcourt_id}_{restaurant_id}_{item_id}_(.+)$"
    match2 = re.search(pattern2, name_without_ext)
    if match2:
        item_name = match2.group(1)
        # Remove step suffix if present (enrich_data, preprocessing, etc.)
        item_name = re.sub(r'_(enrich_data|preprocessing|postprocessing|model_generation.*)$', '', item_name)
        return item_name.replace('_', ' ')
    
    return None


def get_all_names(foodcourt_id: str, restaurant_id: str, item_id: Optional[str] = None, 
                  file_name: Optional[str] = None) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """
    Get foodcourt name, restaurant name, and item name.
    
    Args:
        foodcourt_id: Foodcourt ID
        restaurant_id: Restaurant ID
        item_id: Optional item ID
        file_name: Optional file name to extract item name from
    
    Returns:
        Tuple of (foodcourt_name, restaurant_name, item_name)
    """
    # Get foodcourt and restaurant names from MongoDB
    foodcourt_name, restaurant_name = get_mongo_names(foodcourt_id, restaurant_id)
    
    # Get item name from file name if provided
    item_name = None
    if item_id and file_name:
        item_name = get_item_name_from_filename(file_name, foodcourt_id, restaurant_id, item_id)
    
    return foodcourt_name, restaurant_name, item_name


# Path functions moved to path_utils.py - imported above


def _load_name_mapping() -> Dict[str, Any]:
    """
    Load name_mapping.json to resolve names to IDs.
    Tries dashboard/name_mapping.json first, then root name_mapping.json.
    """
    # Try dashboard/name_mapping.json first (most common location)
    dashboard_mapping = Path(__file__).parent.parent.parent / "dashboard" / "name_mapping.json"
    root_mapping = Path(__file__).parent.parent.parent / "name_mapping.json"
    
    mapping_path = dashboard_mapping if dashboard_mapping.exists() else root_mapping
    
    if not mapping_path.exists():
        LOGGER.warning(f"name_mapping.json not found at {mapping_path}, name resolution will be skipped")
        return {}
    
    try:
        with open(mapping_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as exc:
        LOGGER.warning(f"Error loading name_mapping.json: {exc}, name resolution will be skipped")
        return {}


def _resolve_name_to_id(name: str, mapping_type: str, name_mapping: Dict[str, Any]) -> Optional[str]:
    """
    Resolve a name to an ID using name_mapping.json.
    
    Args:
        name: Name to resolve (foodcourt name or restaurant name)
        mapping_type: "foodcourt" or "restaurant"
        name_mapping: Loaded name_mapping.json dictionary
    
    Returns:
        ID if found, None otherwise
    """
    if not name_mapping:
        return None
    
    # Use direct name_to_id mappings if available (faster lookup)
    if mapping_type == "foodcourt":
        name_to_id = name_mapping.get("foodcourt_name_to_id", {})
        if name_to_id:
            # Direct lookup (case-insensitive)
            name_lower = name.lower().strip()
            for mapped_name, id_val in name_to_id.items():
                if mapped_name and str(mapped_name).lower().strip() == name_lower:
                    return id_val
            # Partial match
            for mapped_name, id_val in name_to_id.items():
                if mapped_name and name_lower in str(mapped_name).lower():
                    LOGGER.info(f"Found partial match for '{name}': '{mapped_name}' (ID: {id_val})")
                    return id_val
        # Fallback to reverse lookup from id_to_name
        id_to_name = name_mapping.get("foodcourt_id_to_name", {})
    elif mapping_type == "restaurant":
        name_to_id = name_mapping.get("restaurant_name_to_id", {})
        if name_to_id:
            # Direct lookup (case-insensitive)
            name_lower = name.lower().strip()
            for mapped_name, id_val in name_to_id.items():
                if mapped_name and str(mapped_name).lower().strip() == name_lower:
                    return id_val
            # Partial match
            for mapped_name, id_val in name_to_id.items():
                if mapped_name and name_lower in str(mapped_name).lower():
                    LOGGER.info(f"Found partial match for '{name}': '{mapped_name}' (ID: {id_val})")
                    return id_val
        # Fallback to reverse lookup from id_to_name
        id_to_name = name_mapping.get("restaurant_id_to_name", {})
    else:
        return None
    
    # Fallback: Search for exact match (case-insensitive) in reverse mapping
    name_lower = name.lower().strip()
    for id_val, mapped_name in id_to_name.items():
        if mapped_name and str(mapped_name).lower().strip() == name_lower:
            return id_val
    
    # Fallback: Search for partial match (contains)
    for id_val, mapped_name in id_to_name.items():
        if mapped_name and name_lower in str(mapped_name).lower():
            LOGGER.info(f"Found partial match for '{name}': '{mapped_name}' (ID: {id_val})")
            return id_val
    
    return None


def _resolve_names_to_ids(step_config: Dict[str, Any], name_mapping: Dict[str, Any]) -> Dict[str, Any]:
    """
    Resolve foodcourt_names to IDs and normalize restaurant_ids and item_ids format.
    
    Note: Restaurant names are NOT resolved because restaurants can have the same name
    across different foodcourts. Use foodcourt_id + restaurant_id pairs instead.
    
    Note: Item names are NOT resolved because items can have the same name across different
    restaurants. Use foodcourt_id + restaurant_id + item_name or item_id instead.
    
    Args:
        step_config: Step configuration dict (may contain foodcourt_ids, foodcourt_names, restaurant_ids, item_ids)
        name_mapping: Loaded name_mapping.json dictionary
    
    Returns:
        Dict with:
        - foodcourt_ids: List of foodcourt IDs (names resolved)
        - restaurant_ids: List that can contain:
          - Simple strings (restaurant IDs) - for backward compatibility
          - Dicts with {foodcourt_id, restaurant_id} - for new format
        - item_ids: List of dicts with {foodcourt_id, restaurant_id, item_name} OR {foodcourt_id, restaurant_id, item_id}
    """
    foodcourt_ids = list(step_config.get("foodcourt_ids", []) or step_config.get("food_court_ids", []))
    restaurant_ids_raw = step_config.get("restaurant_ids", [])
    item_ids_raw = step_config.get("item_ids", [])
    item_names_raw = step_config.get("item_names", [])  # Preserve item_names
    
    # Resolve foodcourt_names to IDs
    foodcourt_names = step_config.get("foodcourt_names", [])
    if foodcourt_names:
        LOGGER.info(f"Resolving {len(foodcourt_names)} foodcourt names to IDs...")
        for name in foodcourt_names:
            resolved_id = _resolve_name_to_id(name, "foodcourt", name_mapping)
            if resolved_id:
                if resolved_id not in foodcourt_ids:
                    foodcourt_ids.append(resolved_id)
                    LOGGER.info(f"  ✓ '{name}' -> {resolved_id}")
            else:
                LOGGER.warning(f"  ✗ Could not resolve foodcourt name: '{name}'")
    
    # Normalize restaurant_ids: support both simple strings and objects with {foodcourt_id, restaurant_id}
    restaurant_ids = []
    for entry in restaurant_ids_raw:
        if isinstance(entry, dict):
            # New format: {foodcourt_id: "...", restaurant_id: "..."}
            if "foodcourt_id" in entry and "restaurant_id" in entry:
                restaurant_ids.append(entry)
            else:
                LOGGER.warning(f"Invalid restaurant entry format: {entry}. Expected {{foodcourt_id, restaurant_id}}")
        elif isinstance(entry, str):
            # Old format: simple restaurant ID string (backward compatibility)
            restaurant_ids.append(entry)
        else:
            LOGGER.warning(f"Invalid restaurant entry type: {type(entry)}. Expected string or dict.")
    
    # Normalize item_ids: support objects with {foodcourt_id, restaurant_id, item_name} OR {foodcourt_id, restaurant_id, item_id}
    item_ids = []
    for entry in item_ids_raw:
        if isinstance(entry, dict):
            # Must have foodcourt_id, restaurant_id, and either item_name or item_id
            if "foodcourt_id" in entry and "restaurant_id" in entry:
                if "item_name" in entry or "item_id" in entry:
                    item_ids.append(entry)
                else:
                    LOGGER.warning(f"Invalid item entry format: {entry}. Expected {{foodcourt_id, restaurant_id, item_name}} or {{foodcourt_id, restaurant_id, item_id}}")
            else:
                LOGGER.warning(f"Invalid item entry format: {entry}. Expected {{foodcourt_id, restaurant_id, item_name}} or {{foodcourt_id, restaurant_id, item_id}}")
        else:
            LOGGER.warning(f"Invalid item entry type: {type(entry)}. Expected dict with {{foodcourt_id, restaurant_id, item_name}} or {{foodcourt_id, restaurant_id, item_id}}")
    
    # Preserve item_names as-is (they are already in the correct format)
    item_names = list(item_names_raw) if item_names_raw else []
    
    return {
        "foodcourt_ids": foodcourt_ids,
        "restaurant_ids": restaurant_ids,
        "item_ids": item_ids,
        "item_names": item_names
    }


def load_retrain_config() -> Dict[str, Any]:
    """
    Load retrain.json configuration.
    Supports both old format (list) and new format (dict with foodcourt_ids and restaurant_ids).
    Now also supports foodcourt_names which will be resolved to IDs using name_mapping.json.
    
    For restaurants: Since restaurant names can be duplicated across foodcourts, you must specify
    both foodcourt_id and restaurant_id. Use the format:
    {
      "restaurant_ids": [
        {"foodcourt_id": "...", "restaurant_id": "..."},
        {"foodcourt_id": "...", "restaurant_id": "..."}
      ]
    }
    
    Or use simple restaurant ID strings for backward compatibility (when foodcourt is already in foodcourt_ids).
    
    Example:
    {
      "model_generation": {
        "foodcourt_names": ["CG-EPIP Kakadu-BLR", "CG-Campus-BLR"],
        "restaurant_ids": [
          {"foodcourt_id": "62720dcd1410b643ccc5eaf0", "restaurant_id": "62b96ba6a3b4411d5b0b508a"},
          {"foodcourt_id": "62720dcd1410b643ccc5eaf0", "restaurant_id": "63085a2195355f366f7fa0f6"}
        ]
      }
    }
    """
    retrain_path = get_retrain_path()
    if not retrain_path.exists():
        LOGGER.warning("retrain.json not found, using empty config")
        return {
            "data_fetch": {"foodcourt_ids": [], "restaurant_ids": [], "item_ids": [], "item_names": []},
            "enrich_data": {"foodcourt_ids": [], "restaurant_ids": [], "item_ids": [], "item_names": []},
            "preprocessing": {"foodcourt_ids": [], "restaurant_ids": [], "item_ids": [], "item_names": []},
            "model_generation": {"foodcourt_ids": [], "restaurant_ids": [], "item_ids": [], "item_names": []},
            "postprocessing": {"foodcourt_ids": [], "restaurant_ids": [], "item_ids": [], "item_names": []},
            "compiled_result_generation": {"foodcourt_ids": [], "restaurant_ids": [], "item_ids": [], "item_names": []}
        }
    
    # Load name_mapping for name resolution
    name_mapping = _load_name_mapping()
    
    try:
        with open(retrain_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
            # Normalize to new format: ensure all steps are dicts with foodcourt_ids and restaurant_ids
            normalized = {}
            for step_name in ["data_fetch", "enrich_data", "preprocessing", "model_generation", 
                            "postprocessing", "compiled_result_generation"]:
                step_config = config.get(step_name, {})
                if isinstance(step_config, list):
                    # Old format: convert list to dict
                    foodcourt_ids = []
                    restaurant_ids = []
                    for entry in step_config:
                        if isinstance(entry, str):
                            foodcourt_ids.append(entry)
                        elif isinstance(entry, dict):
                            if "foodcourt_id" in entry:
                                foodcourt_ids.append(entry["foodcourt_id"])
                            if "restaurant_id" in entry:
                                restaurant_ids.append(entry["restaurant_id"])
                    normalized[step_name] = {
                        "foodcourt_ids": foodcourt_ids,
                        "restaurant_ids": restaurant_ids,
                        "item_ids": [],
                        "item_names": []
                    }
                elif isinstance(step_config, dict):
                    # New format: resolve names to IDs if present
                    resolved = _resolve_names_to_ids(step_config, name_mapping)
                    normalized[step_name] = resolved
                else:
                    normalized[step_name] = {"foodcourt_ids": [], "restaurant_ids": [], "item_ids": [], "item_names": []}
            return normalized
    except Exception as exc:
        LOGGER.error(f"Error loading retrain.json: {exc}")
        return {
            "data_fetch": {"foodcourt_ids": [], "restaurant_ids": [], "item_ids": [], "item_names": []},
            "enrich_data": {"foodcourt_ids": [], "restaurant_ids": [], "item_ids": [], "item_names": []},
            "preprocessing": {"foodcourt_ids": [], "restaurant_ids": [], "item_ids": [], "item_names": []},
            "model_generation": {"foodcourt_ids": [], "restaurant_ids": [], "item_ids": [], "item_names": []},
            "postprocessing": {"foodcourt_ids": [], "restaurant_ids": [], "item_ids": [], "item_names": []},
            "compiled_result_generation": {"foodcourt_ids": [], "restaurant_ids": [], "item_ids": [], "item_names": []}
        }


def is_retrain_config_empty() -> bool:
    """
    Check if retrain.json is empty (all foodcourt_ids, restaurant_ids, and foodcourt_names are empty).
    Note: This checks the raw config before name resolution, so it will detect if names are specified.
    """
    retrain_path = get_retrain_path()
    if not retrain_path.exists():
        return True
    
    try:
        with open(retrain_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
            # Check if all step configs are empty
            for step_name in ["data_fetch", "enrich_data", "preprocessing", "model_generation", 
                            "postprocessing", "compiled_result_generation"]:
                step_config = config.get(step_name, {})
                if isinstance(step_config, list) and len(step_config) > 0:
                    return False
                elif isinstance(step_config, dict):
                    # Check for IDs or names
                    foodcourt_ids = step_config.get("foodcourt_ids", []) or step_config.get("food_court_ids", [])
                    restaurant_ids = step_config.get("restaurant_ids", [])
                    item_ids = step_config.get("item_ids", [])
                    item_names = step_config.get("item_names", [])
                    foodcourt_names = step_config.get("foodcourt_names", [])
                    if (len(foodcourt_ids) > 0 or len(restaurant_ids) > 0 or len(item_ids) > 0 or 
                        len(item_names) > 0 or len(foodcourt_names) > 0):
                        return False
            return True
    except Exception:
        return True


def should_force_retrain(step_name: str, foodcourt_id: str, restaurant_id: Optional[str] = None, 
                         item_name: Optional[str] = None, item_id: Optional[str] = None) -> bool:
    """
    Check if we should force retrain based on retrain.json.
    
    New format: step_config is a dict with "foodcourt_ids", "restaurant_ids", and "item_ids" lists.
    restaurant_ids can contain:
    - Simple strings (restaurant IDs) - matches if foodcourt_id is in foodcourt_ids
    - Dicts with {foodcourt_id, restaurant_id} - matches both foodcourt and restaurant
    
    item_ids can contain:
    - Dicts with {foodcourt_id, restaurant_id, item_name} - matches by name
    - Dicts with {foodcourt_id, restaurant_id, item_id} - matches by ID
    
    Returns True if:
    - item_ids is specified and item matches (most specific check)
    - OR restaurant_id matches a {foodcourt_id, restaurant_id} pair in restaurant_ids
    - OR foodcourt_id is in foodcourt_ids AND (restaurant_ids is empty OR restaurant matches)
    """
    retrain_config = load_retrain_config()
    step_config = retrain_config.get(step_name, {})
    
    if not step_config:
        return False
    
    # Handle new format (dict with foodcourt_ids, restaurant_ids, and item_ids)
    if isinstance(step_config, dict):
        foodcourt_ids = step_config.get("foodcourt_ids", [])
        restaurant_ids = step_config.get("restaurant_ids", [])
        item_ids = step_config.get("item_ids", [])
        
        # First check item_ids if provided (most specific check)
        if item_ids and (item_name or item_id) and restaurant_id:
            for entry in item_ids:
                if isinstance(entry, dict):
                    entry_fc_id = entry.get("foodcourt_id")
                    entry_rest_id = entry.get("restaurant_id")
                    
                    # Must match foodcourt and restaurant first
                    if entry_fc_id == foodcourt_id and entry_rest_id == restaurant_id:
                        # Check if it's using item_name
                        if "item_name" in entry:
                            if item_name and str(entry.get("item_name", "")).strip().lower() == str(item_name).strip().lower():
                                return True
                        # Check if it's using item_id
                        elif "item_id" in entry:
                            if item_id and str(entry.get("item_id", "")).strip() == str(item_id).strip():
                                return True
        
        # If item_ids is specified but item didn't match, return False (strict filtering)
        if item_ids:
            return False
        
        # If restaurant_id is provided, check restaurant_ids (less specific than item_ids)
        if restaurant_id and restaurant_ids:
            for entry in restaurant_ids:
                if isinstance(entry, dict):
                    # New format: {foodcourt_id, restaurant_id} pair
                    if (entry.get("foodcourt_id") == foodcourt_id and 
                        entry.get("restaurant_id") == restaurant_id):
                        return True
                elif isinstance(entry, str):
                    # Old format: simple restaurant ID string
                    # Match if restaurant_id matches AND foodcourt_id is in foodcourt_ids
                    if entry == restaurant_id and foodcourt_id in foodcourt_ids:
                        return True
        
        # If restaurant_ids is specified but restaurant didn't match, return False (strict filtering)
        if restaurant_ids:
            return False
        
        # Check if foodcourt_id is in the list (least specific)
        if foodcourt_id in foodcourt_ids:
            # If restaurant_id is provided but restaurant_ids is empty, process all restaurants in the foodcourt
            if restaurant_id:
                # If restaurant_ids list is empty, process all restaurants in the foodcourt
                if len(restaurant_ids) == 0:
                    return True
                # If we get here, restaurant_ids is not empty but didn't match above, so return False
                return False
            else:
                # Only foodcourt_id provided, match any restaurant
                return True
    
    # Handle old format (list) for backward compatibility
    elif isinstance(step_config, list):
        for entry in step_config:
            if isinstance(entry, str):
                if entry == foodcourt_id:
                    return True
            elif isinstance(entry, dict):
                if entry.get("foodcourt_id") == foodcourt_id:
                    if restaurant_id is None or entry.get("restaurant_id") == restaurant_id:
                        return True
    
    return False


def get_retrain_config_for_step(step_name: str) -> Dict[str, List[str]]:
    """
    Get retrain configuration for a specific step.
    Returns dict with 'foodcourt_ids', 'restaurant_ids', 'item_ids', and 'item_names' lists.
    """
    retrain_config = load_retrain_config()
    step_config = retrain_config.get(step_name, {})
    
    if isinstance(step_config, dict):
        return {
            "foodcourt_ids": step_config.get("foodcourt_ids", []),
            "restaurant_ids": step_config.get("restaurant_ids", []),
            "item_ids": step_config.get("item_ids", []),
            "item_names": step_config.get("item_names", [])
        }
    else:
        # Old format or empty
        return {"foodcourt_ids": [], "restaurant_ids": [], "item_ids": [], "item_names": []}


def matches_item_filter(foodcourt_id: str, restaurant_id: str, item_name: Optional[str] = None,
                       item_id: Optional[str] = None, item_ids_filter: List[Dict[str, Any]] = None,
                       item_names_filter: List[Any] = None) -> bool:
    """
    Check if an item matches the item_ids or item_names filter from retrain.json.
    
    Args:
        foodcourt_id: Foodcourt ID
        restaurant_id: Restaurant ID
        item_name: Optional item name
        item_id: Optional item ID
        item_ids_filter: List of item filter dicts from retrain.json.
                        Each dict can have {foodcourt_id, restaurant_id, item_name} OR
                        {foodcourt_id, restaurant_id, item_id}
        item_names_filter: List of item names. Can be:
                          - List of strings: simple item names (matches any restaurant)
                          - List of dicts: {foodcourt_id, restaurant_id, item_name} (matches specific restaurant)
    
    Returns:
        True if no filters specified OR if item matches a filter entry.
        False if filters are specified but item doesn't match.
    """
    # Check item_names filter
    item_name_matched = False
    if item_names_filter and len(item_names_filter) > 0:
        # Check if item_names_filter contains objects (like item_ids) or simple strings
        first_item = item_names_filter[0] if item_names_filter else None
        is_object_format = isinstance(first_item, dict)
        
        if is_object_format:
            # Format: [{foodcourt_id, restaurant_id, item_name}, ...]
            for entry in item_names_filter:
                if not isinstance(entry, dict):
                    continue
                
                entry_fc_id = str(entry.get("foodcourt_id", "")).strip()
                entry_rest_id = str(entry.get("restaurant_id", "")).strip()
                entry_item_name = str(entry.get("item_name", "")).strip().lower()
                
                # Must match foodcourt, restaurant, and item name
                if (entry_fc_id == str(foodcourt_id).strip() and
                    entry_rest_id == str(restaurant_id).strip() and
                    item_name and str(item_name).strip().lower() == entry_item_name):
                    item_name_matched = True
                    break
        else:
            # Format: ["item name 1", "item name 2", ...] - simple array of strings
            if item_name:
                item_name_lower = str(item_name).strip().lower()
                for filter_name in item_names_filter:
                    if str(filter_name).strip().lower() == item_name_lower:
                        item_name_matched = True
                        break
        
        # If item_names_filter is specified but item doesn't match, check item_ids_filter
        # If both are specified, item must match at least one
        if item_name_matched:
            return True
    
    # Check item_ids filter (more specific, includes foodcourt/restaurant matching)
    if not item_ids_filter or len(item_ids_filter) == 0:
        # No item_ids filter
        # If item_names_filter was checked and matched, we already returned True
        # If item_names_filter was checked and didn't match, return False
        # If no item_names_filter, return True (no filter = allow all)
        if item_names_filter and len(item_names_filter) > 0:
            return False  # item_names_filter was checked but didn't match
        return True  # No filters at all, allow all
    
    # Check each filter entry in item_ids_filter
    for entry in item_ids_filter:
        if not isinstance(entry, dict):
            continue
        
        entry_fc_id = entry.get("foodcourt_id")
        entry_rest_id = entry.get("restaurant_id")
        
        # Must match foodcourt and restaurant first
        if str(entry_fc_id).strip() != str(foodcourt_id).strip() or \
           str(entry_rest_id).strip() != str(restaurant_id).strip():
            continue
        
        # Check if it's using item_name
        if "item_name" in entry:
            entry_item_name = str(entry.get("item_name", "")).strip().lower()
            if item_name and str(item_name).strip().lower() == entry_item_name:
                return True
        
        # Check if it's using item_id
        elif "item_id" in entry:
            entry_item_id = str(entry.get("item_id", "")).strip()
            if item_id and str(item_id).strip() == entry_item_id:
                return True
    
    # No match found in item_ids_filter
    # If item_names_filter was also checked and matched, we would have returned True already
    # If item_names_filter was checked and didn't match, return False
    # If only item_ids_filter was checked, return False
    if item_names_filter and len(item_names_filter) > 0 and not item_name_matched:
        return False  # Both filters checked, neither matched
    return False  # item_ids_filter checked, didn't match


# Path and file name functions moved to path_utils.py - imported above


def clip_text(text: str, max_length: int = 30) -> str:
    """Clip text to max_length and add ellipsis if needed."""
    if not text:
        return ""
    try:
        text_str = str(text)
        if len(text_str) <= max_length:
            return text_str
        return text_str[:max_length-3] + "..."
    except (UnicodeEncodeError, UnicodeDecodeError):
        # If encoding fails, return a safe ASCII representation
        try:
            safe_text = text.encode('utf-8', errors='replace').decode('utf-8')
            if len(safe_text) <= max_length:
                return safe_text
            return safe_text[:max_length-3] + "..."
        except:
            return repr(text)[:max_length-3] + "..."


def is_error_reason(value: str) -> bool:
    """
    Check if a value in file_locator is an error reason (starts with "ERROR:").
    
    Args:
        value: Value from file_locator column
    
    Returns:
        True if value is an error reason, False if it's a file path/hyperlink
    """
    if not value or not isinstance(value, str):
        return False
    value_str = str(value).strip()
    return value_str.startswith("ERROR:")


def is_file_path(value: str) -> bool:
    """
    Check if a value in file_locator is a file path (hyperlink).
    
    Args:
        value: Value from file_locator column
    
    Returns:
        True if value is a file path/hyperlink, False if it's an error reason
    """
    if not value or not isinstance(value, str):
        return False
    value_str = str(value).strip()
    return value_str.startswith("=HYPERLINK")


def has_upstream_error(foodcourt_id: str, restaurant_id: str, item_id: str, item_name: str, 
                       current_step: str) -> bool:
    """
    Check if any upstream step has an error for this FRI combination.
    
    Upstream steps:
    - For model_generation: check enrich_data, preprocessing
    - For postprocessing: check enrich_data, preprocessing, model_generation
    
    Only returns True if there's an explicit ERROR in file_locator.
    If no entry exists, returns False (item may be new or file_locator not updated yet).
    
    Args:
        foodcourt_id, restaurant_id, item_id, item_name: FRI identifiers
        current_step: Current step name ("model_generation" or "postprocessing")
    
    Returns:
        True if any upstream step has an error (ERROR: prefix), False otherwise
    """
    file_locator = get_file_locator()
    
    # Load existing data if not already loaded
    if not file_locator.existing_data and file_locator.locator_path.exists():
        file_locator._load_existing_data()
    
    # Check all entries (existing + new)
    all_entries = file_locator.existing_data + file_locator.data
    
    # Find entry for this FRI
    matching_entry = None
    for entry in all_entries:
        entry_fc_id = str(entry.get("Foodcourt_ID", ""))
        entry_rest_id = str(entry.get("Restaurant_ID", ""))
        entry_item_id = str(entry.get("Item_ID", ""))
        entry_item_name = str(entry.get("Item_Name", ""))
        
        # Match foodcourt and restaurant
        if entry_fc_id == str(foodcourt_id) and entry_rest_id == str(restaurant_id):
            # Match by item_id or item_name
            if (item_id and entry_item_id and str(item_id) == entry_item_id) or \
               (item_name and entry_item_name and str(item_name).lower() == entry_item_name.lower()):
                matching_entry = entry
                break
    
    # If no matching entry found, return False (item may be new, process it)
    if not matching_entry:
        return False
    
    # Check upstream steps based on current step
    if current_step == "model_generation":
        # Check enrich_data and preprocessing
        upstream_steps = ["enrich_data", "preprocessing"]
    elif current_step == "postprocessing":
        # Check enrich_data, preprocessing
        upstream_steps = ["enrich_data", "preprocessing"]
        # Also check model files
        model_files = ["XGBoost_File", "MovingAverage_File"]
        model_status = "model_generation_Status"
    else:
        return False
    
    # Check upstream step File columns for ERROR prefix
    for step in upstream_steps:
        step_col = f"{step}_File"
        if step_col in matching_entry:
            val = str(matching_entry[step_col])
            if is_error_reason(val):
                return True
    
    # For postprocessing, also check model files and status
    if current_step == "postprocessing":
        # Check if both model files have errors or model_generation_Status exists
        both_models_error = True
        for model_col in model_files:
            if model_col in matching_entry:
                val = str(matching_entry[model_col])
                if not is_error_reason(val):
                    both_models_error = False
                    break
            else:
                both_models_error = False
                break
        
        if both_models_error or (model_status in matching_entry and matching_entry[model_status]):
            return True
    
    return False


def create_excel_hyperlink(file_path: Path, display_text: str) -> str:
    """Create Excel hyperlink formula."""
    # Convert to absolute path
    abs_path = file_path.resolve()
    # Normalize path for Windows
    path_str = str(abs_path).replace('\\', '/')
    # Create file URI
    if path_str.startswith('//'):
        uri = f"file:{path_str}"
    else:
        if not path_str.startswith('/'):
            uri = f"file:///{path_str}"
        else:
            uri = f"file://{path_str}"
    # URL encode spaces
    uri = uri.replace(' ', '%20')
    return f'=HYPERLINK("{uri}", "{display_text}")'


# Global instances (will be initialized by run_pipeline.py)
_file_locator: Optional['FileLocator'] = None
_pipeline_logger: Optional['PipelineLogger'] = None
_pipeline_type: str = "FRID_LEVEL"  # Default pipeline type (will be overridden by pipeline_hyperparameters.json)
_pipeline_start_time: Optional[float] = None  # Global pipeline start time


def set_pipeline_type(pipeline_type: str):
    """Set the pipeline type (e.g., 'FRID_LEVEL', 'FR_LEVEL', 'I_LEVEL')."""
    global _pipeline_type
    _pipeline_type = pipeline_type
    # Reinitialize logger with new pipeline type if it exists
    global _pipeline_logger
    if _pipeline_logger is not None:
        _pipeline_logger.pipeline_type = pipeline_type
        _pipeline_logger.log_path = _pipeline_logger.base_dir / "logs" / pipeline_type / "pipeline_logs.xlsx"


def get_pipeline_type() -> str:
    """Get the current pipeline type."""
    return _pipeline_type


def set_pipeline_start_time(start_time: float):
    """Set the global pipeline start time."""
    global _pipeline_start_time
    _pipeline_start_time = start_time


def get_pipeline_start_time() -> Optional[float]:
    """Get the global pipeline start time."""
    return _pipeline_start_time


def get_pipeline_log_dir(step_name: Optional[str] = None) -> Path:
    """
    Get the log directory for the current pipeline.
    
    Args:
        step_name: Optional step name (e.g., 'preprocessing') to create step-specific log dir
    
    Returns:
        Path to the log directory (e.g., output_data/logs/FRID_LEVEL or output_data/logs/FRID_LEVEL/preprocessing)
    """
    base_dir = get_output_base_dir()
    pipeline_type = get_pipeline_type()
    log_dir = base_dir / "logs" / pipeline_type
    if step_name:
        log_dir = log_dir / step_name
    return log_dir


def get_file_locator() -> 'FileLocator':
    """Get the global file locator instance."""
    global _file_locator
    if _file_locator is None:
        _file_locator = FileLocator()
    return _file_locator


def get_pipeline_logger() -> 'PipelineLogger':
    """Get the global pipeline logger instance."""
    global _pipeline_logger
    if _pipeline_logger is None:
        _pipeline_logger = PipelineLogger()
    return _pipeline_logger


class FileLocator:
    """Manages the file_locator.xlsx file."""
    
    def __init__(self):
        self.base_dir = get_output_base_dir()
        self.locator_path = self.base_dir / "file_locator.csv"  # CSV format
        # Removed temp_file_locator directory - using restaurant_tracking instead
        self.data: List[Dict[str, Any]] = []
        self.existing_data: List[Dict[str, Any]] = []
        self.executed_steps: set = set()  # Track which steps were actually executed (not skipped)
        self._load_existing_data()
    
    def _load_existing_data(self):
        """Load existing data from file_locator.csv if it exists."""
        # Try CSV first (new format), then Excel (backward compatibility)
        csv_path = self.locator_path.with_suffix('.csv')
        xlsx_path = self.locator_path.with_suffix('.xlsx')
        
        if csv_path.exists():
            try:
                df = pd.read_csv(csv_path)
                self.existing_data = df.to_dict('records')
                LOGGER.info(f"Loaded {len(self.existing_data)} existing entries from file_locator.csv")
                # Update locator_path to CSV
                self.locator_path = csv_path
                return
            except Exception as exc:
                LOGGER.warning(f"Could not load existing file_locator.csv: {exc}")
        elif xlsx_path.exists():
            try:
                df = pd.read_excel(xlsx_path, sheet_name='File_Locator', engine='openpyxl')
                self.existing_data = df.to_dict('records')
                LOGGER.info(f"Loaded {len(self.existing_data)} existing entries from file_locator.xlsx (backward compatibility)")
                # Update locator_path to CSV for future saves
                self.locator_path = csv_path
                return
            except Exception as exc:
                LOGGER.warning(f"Could not load existing file_locator.xlsx: {exc}")
        
            self.existing_data = []
    
    def _entry_exists(self, foodcourt_id: str, restaurant_id: str, item_id: str, 
                      step_name: str, file_path: Path) -> bool:
        """Check if an entry already exists in existing_data or new data."""
        file_name = file_path.name
        clipped_file_name = clip_text(file_name, 40)
        
        # Check in existing_data
        for entry in self.existing_data:
            if (entry.get("Foodcourt_ID") == foodcourt_id and
                entry.get("Restaurant_ID") == restaurant_id and
                entry.get("Item_ID") == item_id and
                f"{step_name}_File" in entry):
                # Check if the file name matches (compare both clipped and full names)
                existing_file = entry.get(f"{step_name}_File", "")
                # The stored file name is clipped, so compare with clipped version
                # Also check if the actual file name matches (in case of exact match)
                if (existing_file == clipped_file_name or 
                    existing_file == file_name or
                    file_name.startswith(existing_file.rstrip("..."))):
                    return True
        
        # Check in new data
        for entry in self.data:
            if (entry.get("Foodcourt_ID") == foodcourt_id and
                entry.get("Restaurant_ID") == restaurant_id and
                entry.get("Item_ID") == item_id and
                f"{step_name}_File" in entry):
                existing_file = entry.get(f"{step_name}_File", "")
                if (existing_file == clipped_file_name or 
                    existing_file == file_name or
                    file_name.startswith(existing_file.rstrip("..."))):
                    return True
        
        return False
    
    def mark_step_executed(self, step_name: str):
        """Mark a step as executed (not skipped)."""
        self.executed_steps.add(step_name)
    
    def add_discard_reason(self, foodcourt_id: str, foodcourt_name: str, restaurant_id: str,
                          restaurant_name: str, item_id: str, item_name: str,
                          step_name: str, reason: str, model_name: Optional[str] = None):
        """
        Add or update a discard entry in the locator.
        
        When an item is discarded in a step, save "ERROR:<reason>" in the File column.
        For model_generation, use model_generation_Status column if no models were created.
        
        Args:
            foodcourt_id, foodcourt_name: Foodcourt identifiers
            restaurant_id, restaurant_name: Restaurant identifiers
            item_id, item_name: Item identifiers
            step_name: Step name (e.g., "enrich_data", "preprocessing", "model_generation", "postprocessing")
            reason: Discard reason (e.g., "Beverage/MRP item", "File not found", "Empty dataframe")
            model_name: Optional model name (for model_generation step - if provided, puts reason in model's File column)
        """
        # Mark this step as executed (even if item was discarded)
        if model_name and step_name == "model_generation":
            self.executed_steps.add(f"{step_name}_{model_name}")
        else:
            self.executed_steps.add(step_name)
        
        clipped_fc_name = clip_text(foodcourt_name, 20)
        clipped_rest_name = clip_text(restaurant_name, 20)
        clipped_item_name = clip_text(item_name, 30)
        clipped_reason = clip_text(reason, 45)  # Leave room for "ERROR:" prefix
        
        # Format as "ERROR:<reason>"
        error_reason = f"ERROR:{clipped_reason}"
        
        # Determine column name - put reason directly in File column (or Status for model_generation if no model_name)
        if step_name == "model_generation" and not model_name:
            # For model_generation without specific model, use model_generation_Status column
            file_col = "model_generation_Status"
        elif model_name and step_name == "model_generation":
            # For specific model, put reason in that model's File column
            file_col = f"{model_name}_File"
        else:
            # For other steps, put reason in step's File column
            file_col = f"{step_name}_File"
        
        # Check if this FRI already exists in self.data
        # Match by Foodcourt_ID + Restaurant_ID + (Item_ID OR Item_Name)
        # This handles cases where item_id might be None or item_name
        entry_found = False
        entry_idx = None
        for idx, entry in enumerate(self.data):
            entry_fc_id = str(entry.get("Foodcourt_ID", ""))
            entry_rest_id = str(entry.get("Restaurant_ID", ""))
            entry_item_id = str(entry.get("Item_ID", ""))
            entry_item_name = str(entry.get("Item_Name", ""))
            
            # Match foodcourt and restaurant
            if entry_fc_id == str(foodcourt_id) and entry_rest_id == str(restaurant_id):
                # Match by item_id if both are provided and not empty
                if item_id and entry_item_id and str(item_id) == entry_item_id:
                    entry_found = True
                    entry_idx = idx
                    break
                # Match by item_name if item_id doesn't match or is None
                elif entry_item_name and str(item_name).strip() and entry_item_name.lower() == str(item_name).lower():
                    entry_found = True
                    entry_idx = idx
                    break
        
        # Add to self.data (will be merged with existing data when save() is called)
        new_entry = {
            "Foodcourt_ID": foodcourt_id,
            "Foodcourt_Name": clipped_fc_name,
            "Restaurant_ID": restaurant_id,
            "Restaurant_Name": clipped_rest_name,
            "Item_ID": item_id if item_id else item_name,  # Use item_name if item_id is None
            "Item_Name": clipped_item_name,
            file_col: error_reason
        }
        self.data.append(new_entry)
        LOGGER.debug(f"Added discard entry to file_locator: {foodcourt_id}/{restaurant_id}/{item_id or item_name} for {file_col}: {error_reason}")
        
        # Removed temp file saving - using restaurant_tracking instead
    
    def add_file(self, foodcourt_id: str, foodcourt_name: str, restaurant_id: str, 
                 restaurant_name: str, item_id: str, item_name: str, 
                 step_name: str, file_path: Path, model_name: Optional[str] = None):
        """
        Add or update a file entry in the locator.
        
        Simple logic: Check if FRI (Foodcourt, Restaurant, Item) exists in data.
        - If exists: Update the entry with new file location
        - If not: Append new entry
        
        This method automatically marks the step as executed.
        
        Args:
            model_name: Optional model name (e.g., "XGBoost", "MovingAverage"). 
                       If provided, creates model-specific columns like "{model_name}_File".
                       For model_generation step, this allows multiple models per item.
        """
        # Mark this step as executed (not skipped)
        if model_name and step_name == "model_generation":
            # For model_generation, track by model name
            self.executed_steps.add(f"{step_name}_{model_name}")
        else:
            self.executed_steps.add(step_name)
        
        file_name = file_path.name
        clipped_fc_name = clip_text(foodcourt_name, 20)
        clipped_rest_name = clip_text(restaurant_name, 20)
        clipped_item_name = clip_text(item_name, 30)
        clipped_file_name = clip_text(file_name, 40)
        
        # Only use hyperlink (no separate file name column)
        hyperlink = create_excel_hyperlink(file_path, clipped_file_name)
        
        # Determine column names based on whether model_name is provided
        if model_name and step_name == "model_generation":
            # For model_generation, use model-specific columns (only hyperlink, no file name)
            file_col = f"{model_name}_File"  # This will contain the hyperlink
        else:
            # For other steps, use step_name as before (only hyperlink, no file name)
            file_col = f"{step_name}_File"  # This will contain the hyperlink
        
        # Add to self.data (will be combined at the end)
        new_entry = {
            "Foodcourt_ID": foodcourt_id,
            "Foodcourt_Name": clipped_fc_name,
            "Restaurant_ID": restaurant_id,
            "Restaurant_Name": clipped_rest_name,
            "Item_ID": item_id if item_id else item_name,  # Use item_name if item_id is None
            "Item_Name": clipped_item_name,
            file_col: hyperlink
        }
        self.data.append(new_entry)
        LOGGER.debug(f"Added file entry to file_locator: {foodcourt_id}/{restaurant_id}/{item_id or item_name} for {file_col}")
        
        # Removed temp file saving - using restaurant_tracking instead
    
    def combine_and_save_final(self):
        """
        Save file_locator data from self.data combined with existing data.
        Removed temp file functionality - using restaurant_tracking instead.
        
        This should be called at the end of the pipeline (after postprocessing).
        """
        # Helper function to normalize item_name for consistent matching
        def normalize_item_name(name):
            """Normalize item name for consistent matching (case-insensitive, trimmed)"""
            return str(name).strip().lower() if name else ""
        
        # Start with existing data from original file (if exists)
        all_data_dict = {}
        
        # Index existing data from original file
        for entry in self.existing_data:
            fc_id = str(entry.get("Foodcourt_ID", "")).strip()
            rest_id = str(entry.get("Restaurant_ID", "")).strip()
            item_name = normalize_item_name(entry.get("Item_Name", ""))
            # Use (Foodcourt_ID, Restaurant_ID, Item_Name) as the reference key
            key = (fc_id, rest_id, item_name)
            all_data_dict[key] = entry.copy()
        
        # Merge current step data (self.data)
        for entry in self.data:
            fc_id = str(entry.get("Foodcourt_ID", "")).strip()
            rest_id = str(entry.get("Restaurant_ID", "")).strip()
            item_id = str(entry.get("Item_ID", "")).strip()
            item_name = normalize_item_name(entry.get("Item_Name", ""))
            
            # Use (Foodcourt_ID, Restaurant_ID, Item_Name) as the reference key
            key = (fc_id, rest_id, item_name)
            
            # Find or create entry
            if key in all_data_dict:
                # Update existing entry
                existing_entry = all_data_dict[key]
                
                # Update basic info
                existing_entry["Foodcourt_Name"] = entry.get("Foodcourt_Name", existing_entry.get("Foodcourt_Name", ""))
                existing_entry["Restaurant_Name"] = entry.get("Restaurant_Name", existing_entry.get("Restaurant_Name", ""))
                # Keep the most complete Item_Name (prefer non-empty, longer one)
                existing_item_name = str(existing_entry.get("Item_Name", "")).strip()
                new_item_name = str(entry.get("Item_Name", "")).strip()
                if new_item_name and (not existing_item_name or len(new_item_name) > len(existing_item_name)):
                    existing_entry["Item_Name"] = new_item_name
                
                # Update Item_ID if we have it (prefer non-empty value)
                if item_id and (not existing_entry.get("Item_ID") or existing_entry.get("Item_ID") == ""):
                    existing_entry["Item_ID"] = item_id
                
                # Update step columns (only non-empty ones)
                for col in ["enrich_data_File", "preprocessing_File", "XGBoost_File", 
                        "MovingAverage_File", "model_generation_Status", "postprocessing_File"]:
                    if col in entry and pd.notna(entry[col]) and str(entry[col]).strip():
                        existing_entry[col] = entry[col]
                        
                        # Handle model-specific logic
                        if col in ["XGBoost_File", "MovingAverage_File"]:
                            val = str(entry[col])
                            if val.startswith("=HYPERLINK") and "model_generation_Status" in existing_entry:
                                del existing_entry["model_generation_Status"]
                        elif col == "model_generation_Status":
                            # Clear model File columns if they're hyperlinks
                            for model_col in ["XGBoost_File", "MovingAverage_File"]:
                                if model_col in existing_entry:
                                    existing_val = str(existing_entry[model_col])
                                    if existing_val.startswith("=HYPERLINK"):
                                        del existing_entry[model_col]
            else:
                # Add new entry
                all_data_dict[key] = entry.copy()
        
        # Convert to list
        all_data = list(all_data_dict.values())
        
        if not all_data:
            LOGGER.warning("No file data to save in file_locator")
            return
        
        # Create DataFrame
        df = pd.DataFrame(all_data)
        
        # Ensure all expected columns exist
        expected_step_columns = [
            "Foodcourt_ID", "Foodcourt_Name", "Restaurant_ID", "Restaurant_Name", 
            "Item_ID", "Item_Name",
            "enrich_data_File", "preprocessing_File", "XGBoost_File", 
            "MovingAverage_File", "model_generation_Status", "postprocessing_File"
        ]
        
        for col in expected_step_columns:
            if col not in df.columns:
                df[col] = ""
        
        # Remove old Status columns
        cols_to_remove = []
        for col in df.columns:
            if col.endswith("_Status") and col != "model_generation_Status":
                cols_to_remove.append(col)
        if cols_to_remove:
            df = df.drop(columns=cols_to_remove)
        
        # Reorder columns
        base_cols = ["Foodcourt_ID", "Foodcourt_Name", "Restaurant_ID", "Restaurant_Name", "Item_ID", "Item_Name"]
        step_cols = [col for col in expected_step_columns if col not in base_cols and col in df.columns]
        other_cols = [col for col in df.columns if col not in base_cols + step_cols]
        df = df[base_cols + step_cols + other_cols]
        
        # Save final file_locator.csv
        csv_path = self.locator_path.with_suffix('.csv')
        max_retries = 3
        retry_delay = 1
        saved = False
        
        for attempt in range(max_retries):
            try:
                # Save as CSV
                df.to_csv(csv_path, index=False, encoding='utf-8')
                saved = True
                LOGGER.info(f"Final file_locator saved to {csv_path} (total entries: {len(all_data)})")
                # Update locator_path to CSV
                self.locator_path = csv_path
                break
                
            except PermissionError as e:
                if attempt < max_retries - 1:
                    LOGGER.warning(f"PermissionError saving file_locator (attempt {attempt + 1}/{max_retries}): {e}. Retrying in {retry_delay} seconds...")
                    import time
                    time.sleep(retry_delay)
                    retry_delay *= 2
                else:
                    LOGGER.error(f"Failed to save file_locator after {max_retries} attempts due to PermissionError: {e}")
                    raise
            except Exception as e:
                LOGGER.error(f"Error saving file_locator to {self.locator_path}: {e}")
                raise
        
        if not saved:
            LOGGER.error("Failed to save file_locator - unknown error")
            return
    
    def save(self):
        """
        DEPRECATED: Use combine_and_save_final() at the end instead.
        This method now just clears self.data (data is already saved to temp files by add_file/add_discard_reason).
        """
        # Data is already saved to temp files by add_file/add_discard_reason
        # Clear self.data for next step
        self.data = []


def update_model_location(foodcourt_id: str, foodcourt_name: str, restaurant_id: str,
                         restaurant_name: str, item_id: str, item_name: str,
                         model_name: str, model_path: Path):
    """
    Update model_location.xlsx file for a specific model type.
    Creates/updates Excel file with separate sheets for each model type.
    
    Args:
        foodcourt_id, foodcourt_name: Foodcourt identifiers
        restaurant_id, restaurant_name: Restaurant identifiers
        item_id, item_name: Item identifiers
        model_name: Model type name (e.g., "XGBoost", "MovingAverage")
        model_path: Path to the model file (.pkl)
    """
    from src.util.pipeline_utils import get_output_base_dir, get_pipeline_type, clip_text
    OUTPUT_BASE = get_output_base_dir()
    PIPELINE_TYPE = get_pipeline_type()
    
    # Model location file is in models/{model_name}/ directory
    model_type_dir = OUTPUT_BASE / "trainedModel" / PIPELINE_TYPE / "models" / model_name
    model_type_dir.mkdir(parents=True, exist_ok=True)
    location_file = model_type_dir / "model_location.xlsx"
    
    # Load existing data if file exists
    existing_data = []
    if location_file.exists():
        try:
            df = pd.read_excel(location_file, sheet_name=model_name, engine='openpyxl')
            existing_data = df.to_dict('records')
        except Exception:
            # If sheet doesn't exist or file is corrupted, start fresh
            existing_data = []
    
    # Prepare new entry
    new_entry = {
        "Foodcourt_ID": foodcourt_id,
        "Foodcourt_Name": clip_text(foodcourt_name, 20),
        "Restaurant_ID": restaurant_id,
        "Restaurant_Name": clip_text(restaurant_name, 20),
        "Item_ID": item_id,
        "Item_Name": clip_text(item_name, 30),
        "Model_Name": model_name,
        "Model_Location": str(model_path.resolve())
    }
    
    # Check if entry exists (same FRI + model)
    entry_key = (foodcourt_id, restaurant_id, item_id)
    updated = False
    for idx, entry in enumerate(existing_data):
        if (str(entry.get("Foodcourt_ID", "")) == foodcourt_id and
            str(entry.get("Restaurant_ID", "")) == restaurant_id and
            str(entry.get("Item_ID", "")) == item_id):
            # Update existing entry
            existing_data[idx] = new_entry
            updated = True
            break
    
    if not updated:
        # Append new entry
        existing_data.append(new_entry)
    
    # Save to Excel with sheet named after model type
    df = pd.DataFrame(existing_data)
    with pd.ExcelWriter(location_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=model_name, index=False)
        
        # Auto-adjust column widths
        worksheet = writer.sheets[model_name]
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width


class PipelineLogger:
    """Manages the log Excel file with multiple sheets."""
    
    def __init__(self, pipeline_type: Optional[str] = None):
        self.base_dir = get_output_base_dir()
        self.pipeline_type = pipeline_type or get_pipeline_type()
        self.log_path = self.base_dir / "logs" / self.pipeline_type / "pipeline_logs.xlsx"
        self.logs: Dict[str, List[Dict[str, Any]]] = {
            "enrichment_errors": [],
            "preprocessing_errors": [],
            "postprocessing_errors": [],
            "data_fetch_errors": [],
            "general_errors": [],
            "process_log": [],  # For connection status, warnings, process start/end, results
            "enrichment_logs": [],  # For step2: restaurant/item not processed with input file links
            "preprocessing_logs": [],  # For step3: discard report merged here
            "model_generation_logs": [],  # For step4: items not processed (replaces model_training_errors)
            "postprocessing_logs": [],  # For step5: items not processed
            "compiled_result_logs": []  # For step6: items not processed
        }
    
    def log_enrichment_error(self, foodcourt_id: str, foodcourt_name: str, 
                             restaurant_id: str, restaurant_name: str,
                             item_id: Optional[str], item_name: Optional[str],
                             reason: str):
        """Log an enrichment error."""
        self.logs["enrichment_errors"].append({
            "Foodcourt_ID": foodcourt_id,
            "Foodcourt_Name": clip_text(foodcourt_name, 30),
            "Restaurant_ID": restaurant_id,
            "Restaurant_Name": clip_text(restaurant_name, 30),
            "Item_ID": item_id or "",
            "Item_Name": clip_text(item_name or "", 40),
            "Reason": clip_text(reason, 100)
        })
    
    def log_enrichment_log(self, foodcourt_id: str, foodcourt_name: str,
                           restaurant_id: str, restaurant_name: str,
                           item_id: Optional[str], item_name: Optional[str],
                           input_file_link: str, reason: str):
        """Log enrichment step entry (restaurant/item not processed) with input file hyperlink."""
        self.logs["enrichment_logs"].append({
            "Foodcourt_ID": foodcourt_id,
            "Foodcourt_Name": clip_text(foodcourt_name, 30),
            "Restaurant_ID": restaurant_id,
            "Restaurant_Name": clip_text(restaurant_name, 30),
            "Item_ID": item_id or "",
            "Item_Name": clip_text(item_name or "", 40),
            "Input_File_Link": input_file_link,
            "Reason": clip_text(reason, 100)
        })
    
    def log_preprocessing_log(self, foodcourt_id: str, foodcourt_name: str,
                              restaurant_id: str, restaurant_name: str,
                              item_id: Optional[str], item_name: Optional[str],
                              reason: str):
        """Log preprocessing step entry (discarded items)."""
        self.logs["preprocessing_logs"].append({
            "Foodcourt_ID": foodcourt_id,
            "Foodcourt_Name": clip_text(foodcourt_name, 30),
            "Restaurant_ID": restaurant_id,
            "Restaurant_Name": clip_text(restaurant_name, 30),
            "Item_ID": item_id or "",
            "Item_Name": clip_text(item_name or "", 40),
            "Reason": clip_text(reason, 100)
        })
    
    def log_model_generation_log(self, foodcourt_id: str, foodcourt_name: str,
                                 restaurant_id: str, restaurant_name: str,
                                 item_id: str, item_name: str,
                                 input_file_link: str, reason: str):
        """Log model generation step entry (items not processed) with input file hyperlink."""
        self.logs["model_generation_logs"].append({
            "Foodcourt_ID": foodcourt_id,
            "Foodcourt_Name": clip_text(foodcourt_name, 30),
            "Restaurant_ID": restaurant_id,
            "Restaurant_Name": clip_text(restaurant_name, 30),
            "Item_ID": item_id,
            "Item_Name": clip_text(item_name, 40),
            "Input_File_Link": input_file_link,
            "Reason": clip_text(reason, 100)
        })
    
    def log_postprocessing_log(self, foodcourt_id: str, foodcourt_name: str,
                               restaurant_id: str, restaurant_name: str,
                               item_id: str, item_name: str,
                               input_file_link: str, reason: str):
        """Log postprocessing step entry (items not processed) with input file hyperlink."""
        self.logs["postprocessing_logs"].append({
            "Foodcourt_ID": foodcourt_id,
            "Foodcourt_Name": clip_text(foodcourt_name, 30),
            "Restaurant_ID": restaurant_id,
            "Restaurant_Name": clip_text(restaurant_name, 30),
            "Item_ID": item_id,
            "Item_Name": clip_text(item_name, 40),
            "Input_File_Link": input_file_link,
            "Reason": clip_text(reason, 100)
        })
    
    def log_preprocessing_error(self, foodcourt_id: str, foodcourt_name: str,
                               restaurant_id: str, restaurant_name: str,
                               item_id: Optional[str], item_name: Optional[str],
                               reason: str):
        """Log a preprocessing error."""
        self.logs["preprocessing_errors"].append({
            "Foodcourt_ID": foodcourt_id,
            "Foodcourt_Name": clip_text(foodcourt_name, 30),
            "Restaurant_ID": restaurant_id,
            "Restaurant_Name": clip_text(restaurant_name, 30),
            "Item_ID": item_id or "",
            "Item_Name": clip_text(item_name or "", 40),
            "Reason": clip_text(reason, 100)
        })
    
    def log_postprocessing_error(self, foodcourt_id: str, foodcourt_name: str,
                                 restaurant_id: str, restaurant_name: str,
                                 item_id: str, item_name: str,
                                 reason: str):
        """Log a postprocessing error."""
        self.logs["postprocessing_errors"].append({
            "Foodcourt_ID": foodcourt_id,
            "Foodcourt_Name": clip_text(foodcourt_name, 30),
            "Restaurant_ID": restaurant_id,
            "Restaurant_Name": clip_text(restaurant_name, 30),
            "Item_ID": item_id,
            "Item_Name": clip_text(item_name, 40),
            "Reason": clip_text(reason, 100)
        })
    
    def log_data_fetch_error(self, foodcourt_id: str, foodcourt_name: str,
                            restaurant_id: str, restaurant_name: str,
                            reason: str):
        """Log a data fetch error."""
        self.logs["data_fetch_errors"].append({
            "Foodcourt_ID": foodcourt_id,
            "Foodcourt_Name": clip_text(foodcourt_name, 30),
            "Restaurant_ID": restaurant_id,
            "Restaurant_Name": clip_text(restaurant_name, 30),
            "Reason": clip_text(reason, 100)
        })
    
    def log_general_error(self, step: str, message: str, details: Optional[str] = None):
        """Log a general error."""
        self.logs["general_errors"].append({
            "Step": step,
            "Message": clip_text(message, 100),
            "Details": clip_text(details or "", 200)
        })
    
    def log_connection_status(self, service: str, status: str, details: str = ""):
        """Log connection status (MongoDB, MySQL, etc.) with emoji."""
        emoji = "✅" if status.lower() == "connected" or status.lower() == "success" else "❌"
        message = f"{emoji} {service}"
        if details:
            message += f" ({details})"
        self.logs["process_log"].append({
            "Timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Step": "Connection",
            "Message": message
        })
    
    def log_warning(self, step: str, message: str):
        """Log a warning (even if not shown in console)."""
        self.logs["process_log"].append({
            "Timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Step": step,
            "Message": f"⚠️ WARNING: {message}"
        })
    
    def log_info(self, step: str, message: str):
        """Log an info message."""
        self.logs["process_log"].append({
            "Timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Step": step,
            "Message": message
        })
    
    def log_process_start(self, step_name: str, summary: Dict[str, Any]):
        """Log process start with summary information."""
        self.logs["process_log"].append({
            "Timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Step": step_name,
            "Message": f"🚀 PROCESS STARTED: {step_name}"
        })
        # Add summary details
        for key, value in summary.items():
            self.logs["process_log"].append({
                "Timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Step": step_name,
                "Message": f"  - {key}: {value}"
            })
    
    def log_process_summary(self, step_name: str, total_count: int, processed_count: int, 
                           error_tracking: Dict[str, List[Dict[str, Any]]]):
        """Log comprehensive process summary with error details to a dedicated sheet."""
        sheet_name = f"{step_name}_summary"
        if sheet_name not in self.logs:
            self.logs[sheet_name] = []
        
        # Summary row
        remaining_count = total_count - processed_count
        self.logs[sheet_name].append({
            "Metric": "Total Count",
            "Value": total_count,
            "Details": ""
        })
        self.logs[sheet_name].append({
            "Metric": "Processed Successfully",
            "Value": processed_count,
            "Details": ""
        })
        self.logs[sheet_name].append({
            "Metric": "Remaining/Failed",
            "Value": remaining_count,
            "Details": ""
        })
        self.logs[sheet_name].append({
            "Metric": "",
            "Value": "",
            "Details": ""
        })
        
        # Error breakdown
        self.logs[sheet_name].append({
            "Metric": "Error Category",
            "Value": "Count",
            "Details": "Details"
        })
        
        for error_type, error_list in error_tracking.items():
            if error_list:
                self.logs[sheet_name].append({
                    "Metric": error_type.replace("_", " ").title(),
                    "Value": len(error_list),
                    "Details": f"See details below"
                })
        
        self.logs[sheet_name].append({
            "Metric": "",
            "Value": "",
            "Details": ""
        })
        
        # Detailed error list
        self.logs[sheet_name].append({
            "Metric": "Foodcourt ID",
            "Value": "Restaurant ID",
            "Details": "Reason"
        })
        
        for error_type, error_list in error_tracking.items():
            for error in error_list:
                self.logs[sheet_name].append({
                    "Metric": error.get("foodcourt_id", ""),
                    "Value": error.get("restaurant_id", ""),
                    "Details": error.get("reason", "")
                })
    
    def log_process_results(self, step_name: str, results: Dict[str, Any]):
        """Log process results (foodcourts, restaurants, items processed, etc.)."""
        self.logs["process_log"].append({
            "Timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Step": step_name,
            "Message": f"✅ PROCESS COMPLETED: {step_name}"
        })
        # Add result details
        for key, value in results.items():
            self.logs["process_log"].append({
                "Timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Step": step_name,
                "Message": f"  - {key}: {value}"
            })
    
    def save(self):
        """Save all logs to Excel file."""
        # Create logs directory (pipeline-specific)
        self.log_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Map internal log keys to Excel sheet names
        sheet_name_mapping = {
            "process_log": "Process Log",
            "enrichment_logs": "Enrichment Logs",
            "preprocessing_logs": "Preprocessing Logs",
            "model_generation_logs": "Model Generation Logs",
            "postprocessing_logs": "Postprocessing Logs",
            "compiled_result_logs": "Compiled Result Logs",
            "enrichment_errors": "Enrichment Errors",
            "preprocessing_errors": "Preprocessing Errors",
            "postprocessing_errors": "Postprocessing Errors",
            "data_fetch_errors": "Data Fetch Errors",
            "general_errors": "General Errors"
        }
        
        def safe_str(value):
            """Safely convert value to string, handling Unicode encoding issues."""
            if value is None:
                return ""
            try:
                # First convert to string
                if not isinstance(value, str):
                    value = str(value)
                # Ensure it can be encoded/decoded properly
                # This will raise UnicodeEncodeError if there are problematic characters
                value.encode('utf-8')
                return value
            except (UnicodeEncodeError, UnicodeDecodeError, AttributeError):
                # If encoding fails, return a safe representation
                try:
                    if isinstance(value, str):
                        return value.encode('utf-8', errors='replace').decode('utf-8')
                    else:
                        str_value = str(value)
                        return str_value.encode('utf-8', errors='replace').decode('utf-8')
                except:
                    return repr(value)[:100]  # Limit length to avoid issues
        
        import tempfile
        import shutil
        
        # Try to save to the target file, with fallback to temp file if permission denied
        try:
            with pd.ExcelWriter(self.log_path, engine='openpyxl') as writer:
                sheets_created = 0
                # Save all log sheets
                for log_key, log_data in self.logs.items():
                    if not log_data:
                        continue
                    
                    # Get sheet name from mapping, or use log_key as fallback
                    sheet_name = sheet_name_mapping.get(log_key, log_key.replace("_", " ").title())
                    
                    # Ensure all string values in log_data are properly encoded
                    sanitized_log_data = []
                    for entry in log_data:
                        sanitized_entry = {}
                        for key, value in entry.items():
                            sanitized_entry[key] = safe_str(value) if isinstance(value, str) else value
                        sanitized_log_data.append(sanitized_entry)
                    
                    df = pd.DataFrame(sanitized_log_data)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    sheets_created += 1
                    
                    # Auto-adjust column widths
                    worksheet = writer.sheets[sheet_name]
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                cell_value = cell.value
                                if cell_value is not None:
                                    # Use safe_str to handle Unicode properly
                                    cell_str = safe_str(cell_value)
                                    if len(cell_str) > max_length:
                                        max_length = len(cell_str)
                            except Exception:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Ensure at least one sheet is created (Excel requires at least one visible sheet)
                if sheets_created == 0:
                    # Create an empty "Summary" sheet
                    pd.DataFrame({"Message": ["No logs to display"]}).to_excel(
                        writer, sheet_name="Summary", index=False
                    )
            
            LOGGER.info(f"Pipeline logs saved to {self.log_path}")
        
        except PermissionError:
            # If permission denied (file locked), save to temp file and log warning
            LOGGER.warning(f"Permission denied writing to {self.log_path}. File may be open in Excel.")
            LOGGER.warning("Attempting to save to temporary file...")
            
            try:
                # Save to temp file in same directory
                temp_path = self.log_path.parent / f"{self.log_path.stem}_temp_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
                    # Save all log sheets (same code as above)
                    for log_key, log_data in self.logs.items():
                        if not log_data:
                            continue
                        
                        sheet_name = sheet_name_mapping.get(log_key, log_key.replace("_", " ").title())
                        
                        sanitized_log_data = []
                        for entry in log_data:
                            sanitized_entry = {}
                            for key, value in entry.items():
                                sanitized_entry[key] = safe_str(value) if isinstance(value, str) else value
                            sanitized_log_data.append(sanitized_entry)
                        
                        df = pd.DataFrame(sanitized_log_data)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        worksheet = writer.sheets[sheet_name]
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = get_column_letter(column[0].column)
                            for cell in column:
                                try:
                                    cell_value = cell.value
                                    if cell_value is not None:
                                        cell_str = safe_str(cell_value)
                                        if len(cell_str) > max_length:
                                            max_length = len(cell_str)
                                except Exception:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                
                LOGGER.warning(f"Pipeline logs saved to temporary file: {temp_path}")
                LOGGER.warning("Please close the original file and manually replace it, or use the temp file.")
            except Exception as e:
                LOGGER.error(f"Failed to save pipeline logs even to temporary file: {e}")
        
        except Exception as e:
            LOGGER.error(f"Error saving pipeline logs to {self.log_path}: {e}")


def save_dataframe_to_excel(df: pd.DataFrame, file_path: Path, sheet_name: str = "Sheet1"):
    """
    Save a DataFrame to CSV file (replaces Excel).
    Maintains backward compatibility by accepting sheet_name parameter but saves as CSV.
    """
    file_path.parent.mkdir(parents=True, exist_ok=True)
    # Ensure file_path has .csv extension
    if file_path.suffix.lower() != '.csv':
        file_path = file_path.with_suffix('.csv')
    
    # Convert date columns to string format (YYYY-MM-DD) for CSV
    df_to_save = df.copy()
    for col in df_to_save.columns:
        if df_to_save[col].dtype == 'datetime64[ns]':
            df_to_save[col] = pd.to_datetime(df_to_save[col]).dt.strftime('%Y-%m-%d')
        elif 'date' in col.lower() and df_to_save[col].dtype == 'object':
            # Try to convert date-like strings
            try:
                df_to_save[col] = pd.to_datetime(df_to_save[col]).dt.strftime('%Y-%m-%d')
            except:
                pass
    
    df_to_save.to_csv(file_path, index=False, encoding='utf-8')


def create_result_file_with_sheets(file_path: Path, training_data: pd.DataFrame, 
                                   validation_data: pd.DataFrame, model_name: str):
    """
    Create separate CSV files for Training Data and Validation Data.
    File naming: {base_name}_training.csv and {base_name}_validation.csv
    
    Args:
        file_path: Base path (will be modified to add _training or _validation suffix)
        training_data: DataFrame with training results
        validation_data: DataFrame with validation results
        model_name: Model name (for logging)
    """
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Remove .csv extension if present, then add _training.csv and _validation.csv
    base_path = file_path.with_suffix('')
    if base_path.name.endswith('_training') or base_path.name.endswith('_validation'):
        # Already has suffix, remove it
        base_path = Path(str(base_path).rsplit('_', 1)[0])
    
    training_path = Path(f"{base_path}_training.csv")
    validation_path = Path(f"{base_path}_validation.csv")
    
    # Ensure dates are strings (YYYY-MM-DD format) before saving
    if not training_data.empty:
        training_to_save = training_data.copy()
        if "date" in training_to_save.columns:
            training_to_save["date"] = pd.to_datetime(training_to_save["date"], errors='coerce').dt.strftime('%Y-%m-%d')
        training_to_save.to_csv(training_path, index=False, encoding='utf-8')
        LOGGER.debug(f"Saved training results to {training_path}")
    
    if not validation_data.empty:
        validation_to_save = validation_data.copy()
        if "date" in validation_to_save.columns:
            validation_to_save["date"] = pd.to_datetime(validation_to_save["date"], errors='coerce').dt.strftime('%Y-%m-%d')
        validation_to_save.to_csv(validation_path, index=False, encoding='utf-8')
        LOGGER.debug(f"Saved validation results to {validation_path}")


def validate_output_files(step_name: str, output_dir: Path, foodcourt_id: str, 
                          saved_files: List[Path], min_files_to_check: int = 2) -> bool:
    """
    Validate that output files exist after processing.
    Checks the first min_files_to_check files to ensure they were saved correctly.
    
    Args:
        step_name: Name of the step (for logging)
        output_dir: Base output directory for the step
        foodcourt_id: Foodcourt ID (files are in subdirectory)
        saved_files: List of file paths that should have been saved
        min_files_to_check: Number of files to check (default: 2)
    
    Returns:
        True if all checked files exist, False otherwise
    """
    if not saved_files:
        LOGGER.warning(f"[{step_name}] No files to validate for foodcourt {foodcourt_id}")
        return False
    
    # Check first min_files_to_check files
    files_to_check = saved_files[:min_files_to_check]
    missing_files = []
    
    for file_path in files_to_check:
        if not file_path.exists():
            missing_files.append(str(file_path))
    
    if missing_files:
        LOGGER.error(f"[{step_name}] VALIDATION FAILED: {len(missing_files)} out of {len(files_to_check)} checked files are missing:")
        for missing_file in missing_files:
            LOGGER.error(f"  - Missing: {missing_file}")
        return False
    else:
        # Removed INFO log: "[enrich_data] VALIDATION PASSED: All X checked files exist..."
        return True


def load_pipeline_config() -> Dict[str, Any]:
    """
    Load pipeline configuration from pipeline_hyperparameters.json.
    
    Returns:
        Dictionary containing pipeline configuration with structure:
        {
            "active_pipeline_type": "FRID_LEVEL",
            "pipelines": {
                "FRID_LEVEL": {
                    "description": "...",
                    "steps": {
                        "enrich_data": {
                            "module": "src.step2_data_enrichment",
                            "function": "main",
                            "enabled": true
                        },
                        ...
                    }
                }
            }
        }
    """
    config_path = get_input_base_dir() / "pipeline_hyperparameters.json"
    
    if not config_path.exists():
        # Return default configuration if file doesn't exist
        LOGGER.warning(f"pipeline_hyperparameters.json not found at {config_path}. Using default FRID_LEVEL configuration.")
        return {
            "active_pipeline_type": "FRID_LEVEL",
            "pipelines": {
                "FRID_LEVEL": {
                    "description": "Foodcourt-Restaurant-Item-Date Level Pipeline",
                    "steps": {
                        "data_fetch": {
                            "module": "src.step1_fetch_data",
                            "function": "main",
                            "enabled": False
                        },
                        "enrich_data": {
                            "module": "src.step2_data_enrichment",
                            "function": "main",
                            "enabled": True
                        },
                        "preprocessing": {
                            "module": "src.step3_data_preprocessing",
                            "function": "main",
                            "enabled": True
                        },
                        "model_generation": {
                            "module": "src.step4_model_generation",
                            "function": "main",
                            "enabled": True
                        },
                        "postprocessing": {
                            "module": "src.step5_postprocessing",
                            "function": "main",
                            "enabled": True
                        },
                        "compiled_result_generation": {
                            "module": "src.step6_compiled_result_generation",
                            "function": "main",
                            "enabled": True
                        }
                    }
                }
            }
        }
    
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        return config
    except Exception as exc:
        LOGGER.error(f"Failed to load pipeline configuration from {config_path}: {exc}")
        raise RuntimeError(f"Failed to load pipeline configuration: {exc}") from exc


def get_pipeline_step_config(pipeline_type: Optional[str] = None, step_name: str = "") -> Optional[Dict[str, Any]]:
    """
    Get configuration for a specific step in a pipeline type.
    
    Args:
        pipeline_type: Pipeline type (e.g., "FRID_LEVEL"). If None, uses active pipeline type.
        step_name: Name of the step (e.g., "enrich_data", "preprocessing")
    
    Returns:
        Step configuration dict with keys: module, function, enabled
        Returns None if step not found or disabled
    """
    config = load_pipeline_config()
    
    if pipeline_type is None:
        pipeline_type = config.get("active_pipeline_type", "FRID_LEVEL")
    
    pipelines = config.get("pipelines", {})
    pipeline_config = pipelines.get(pipeline_type, {})
    steps = pipeline_config.get("steps", {})
    
    step_config = steps.get(step_name)
    if step_config and step_config.get("enabled", True):
        return step_config
    
    return None


def get_active_pipeline_type() -> str:
    """
    Get the active pipeline type from pipeline_hyperparameters.json.
    
    Returns:
        Active pipeline type string (e.g., "FRID_LEVEL")
    """
    config = load_pipeline_config()
    return config.get("active_pipeline_type", "FRID_LEVEL")


def import_step_function(step_name: str, pipeline_type: Optional[str] = None):
    """
    Dynamically import and return the main function for a pipeline step.
    
    Args:
        step_name: Name of the step (e.g., "enrich_data", "preprocessing")
        pipeline_type: Pipeline type (e.g., "FRID_LEVEL"). If None, uses active pipeline type.
    
    Returns:
        The main function for the step
    
    Raises:
        RuntimeError: If step configuration not found or import fails
    """
    step_config = get_pipeline_step_config(pipeline_type, step_name)
    
    if step_config is None:
        raise RuntimeError(f"Step '{step_name}' not found or disabled in pipeline configuration")
    
    module_name = step_config.get("module")
    function_name = step_config.get("function", "main")
    
    if not module_name:
        raise RuntimeError(f"Module not specified for step '{step_name}' in pipeline configuration")
    
    try:
        # Dynamically import the module
        import importlib
        module = importlib.import_module(module_name)
        
        # Get the function from the module
        if not hasattr(module, function_name):
            raise RuntimeError(f"Function '{function_name}' not found in module '{module_name}'")
        
        step_function = getattr(module, function_name)
        return step_function
    
    except ImportError as exc:
        raise RuntimeError(f"Failed to import module '{module_name}' for step '{step_name}': {exc}") from exc
    except Exception as exc:
        raise RuntimeError(f"Failed to load step function for '{step_name}': {exc}") from exc

