"""
Generate Excel report from restaurant tracking JSON files.

This script creates a report with 2 sheets:
1. Successful Compilation - Items that completed all steps successfully
2. Remaining Data - Items that failed at any step

Can be run independently at any time to generate reports.
"""

import json
import logging
import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s - %(message)s",
)
LOGGER = logging.getLogger(__name__)


def create_excel_hyperlink(file_path: Path, display_text: str) -> str:
    """Create Excel hyperlink formula."""
    abs_path = file_path.resolve()
    path_str = str(abs_path).replace('\\', '/')
    
    if path_str.startswith('//'):
        uri = f"file:{path_str}"
    else:
        if not path_str.startswith('/'):
            uri = f"file:///{path_str}"
        else:
            uri = f"file://{path_str}"
    
    return f'=HYPERLINK("{uri}", "{display_text}")'


def load_all_tracking_files(tracking_dir: Path) -> List[Dict[str, Any]]:
    """
    Load all restaurant tracking JSON files.
    
    Returns list of tracking data dictionaries.
    """
    tracking_files = []
    
    if not tracking_dir.exists():
        LOGGER.warning(f"Tracking directory not found: {tracking_dir}")
        return tracking_files
    
    # Find all JSON files in subdirectories
    for json_file in tracking_dir.rglob("*.json"):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                tracking_files.append(data)
        except Exception as e:
            LOGGER.warning(f"Failed to load {json_file}: {e}")
    
    LOGGER.info(f"Loaded {len(tracking_files)} tracking files")
    return tracking_files


def is_item_successful(item_data: Dict[str, Any]) -> Tuple[bool, Optional[str]]:
    """
    Check if an item completed all steps successfully.
    
    Returns:
        Tuple of (is_successful, error_reason_if_failed)
    """
    required_steps = ["enrich_data", "preprocessing", "model_generation"]
    
    for step in required_steps:
        if step not in item_data:
            return False, f"Missing step: {step}"
        
        step_data = item_data[step]
        if isinstance(step_data, dict) and step_data.get("error", False):
            error_msg = step_data.get("msg", "Unknown error")
            return False, f"{step}: {error_msg}"
    
    # Check if model_generation has at least one model
    model_gen = item_data.get("model_generation", {})
    if not isinstance(model_gen, dict) or "models" not in model_gen:
        return False, "model_generation: No models found"
    
    models = model_gen.get("models", {})
    if not models:
        return False, "model_generation: No models available"
    
    return True, None


def find_best_model(item_data: Dict[str, Any]) -> Tuple[Optional[str], bool, Dict[str, Any]]:
    """
    Find the best model by comparing all models from postprocessing and model_generation.
    Best model is the one with highest avg_abs_accuracy_pct.
    
    Returns:
        Tuple of (best_model_name, postprocessing_used, metrics_dict)
    """
    postprocessing = item_data.get("postprocessing", {})
    model_generation = item_data.get("model_generation", {})
    
    best_model = None
    best_accuracy = -float('inf')
    postprocessing_used = False
    best_metrics = {}
    
    # Check all postprocessing models
    if postprocessing and not postprocessing.get("error", False):
        postprocessing_models = postprocessing.get("models", {})
        for model_name, model_data in postprocessing_models.items():
            validation = model_data.get("validation", {})
            accuracy = validation.get("avg_abs_accuracy_pct", -float('inf'))
            if accuracy > best_accuracy:
                best_accuracy = accuracy
                best_model = model_name
                postprocessing_used = True
                best_metrics = {
                    "abs_avg_deviation": validation.get("abs_avg_deviation", 0.0),
                    "avg_abs_accuracy_pct": accuracy
                }
    
    # Check all model_generation models
    if model_generation and not model_generation.get("error", False):
        models = model_generation.get("models", {})
        for model_name, model_data in models.items():
            validation = model_data.get("validation", {})
            accuracy = validation.get("avg_abs_accuracy_pct", -float('inf'))
            if accuracy > best_accuracy:
                best_accuracy = accuracy
                best_model = model_name
                postprocessing_used = False
                best_metrics = {
                    "abs_avg_deviation": validation.get("abs_avg_deviation", 0.0),
                    "avg_abs_accuracy_pct": accuracy
                }
    
    if best_model:
        return best_model, postprocessing_used, best_metrics
    
    # Fallback: use compiled_results if available
    compiled_results = item_data.get("compiled_results", {})
    if compiled_results and not compiled_results.get("error", False):
        capped_summary = compiled_results.get("capped_summary", {})
        if capped_summary:
            best_model = capped_summary.get("model", "")
            postprocessing_used = capped_summary.get("postProcessing_used", False)
            abs_avg_deviation = capped_summary.get("abs_avg_deviation", 0.0)
            abs_avg_accuracy = capped_summary.get("abs_avg_accuracy", 0.0)
            
            return best_model, postprocessing_used, {
                "abs_avg_deviation": abs_avg_deviation,
                "avg_abs_accuracy_pct": abs_avg_accuracy
            }
    
    return None, False, {}


def get_file_paths(item_data: Dict[str, Any], best_model: str, postprocessing_used: bool) -> Tuple[Optional[str], Optional[str]]:
    """
    Get training and validation file paths for the best model.
    
    Returns:
        Tuple of (training_file_path, validation_file_path)
    """
    training_path = None
    validation_path = None
    
    if postprocessing_used:
        # Get from model_generation (postprocessing doesn't have file paths)
        model_generation = item_data.get("model_generation", {})
        models = model_generation.get("models", {})
        model_data = models.get(best_model, {})
        
        training_data = model_data.get("training", {})
        validation_data = model_data.get("validation", {})
        
        training_path = training_data.get("file_path")
        validation_path = validation_data.get("file_path")
    else:
        # Get from model_generation
        model_generation = item_data.get("model_generation", {})
        models = model_generation.get("models", {})
        model_data = models.get(best_model, {})
        
        training_data = model_data.get("training", {})
        validation_data = model_data.get("validation", {})
        
        training_path = training_data.get("file_path")
        validation_path = validation_data.get("file_path")
    
    return training_path, validation_path


def get_accuracy_reasons(item_data: Dict[str, Any], best_model: str, postprocessing_used: bool) -> Optional[str]:
    """
    Get accuracy reasons for the best model (only if accuracy < 75%).
    
    Returns:
        Comma-separated string of reasons or None
    """
    accuracy = None
    
    if postprocessing_used:
        postprocessing = item_data.get("postprocessing", {})
        models = postprocessing.get("models", {})
        model_data = models.get(best_model, {})
        validation = model_data.get("validation", {})
        accuracy = validation.get("avg_abs_accuracy_pct", None)
        reasons = validation.get("accuracy_reasons")
    else:
        model_generation = item_data.get("model_generation", {})
        models = model_generation.get("models", {})
        model_data = models.get(best_model, {})
        validation = model_data.get("validation", {})
        accuracy = validation.get("avg_abs_accuracy_pct", None)
        reasons = validation.get("accuracy_reasons")
    
    # Only return reasons if accuracy < 75%
    if accuracy is not None and accuracy < 75.0 and reasons:
        return "; ".join(reasons) if isinstance(reasons, list) else str(reasons)
    
    return None


def get_total_and_active_days(item_data: Dict[str, Any], best_model: str, postprocessing_used: bool) -> Tuple[int, int]:
    """
    Get total_days and active_days from training data (not validation data).
    
    Returns:
        Tuple of (total_days, active_days)
    """
    # Get training data from the best model
    if postprocessing_used:
        postprocessing = item_data.get("postprocessing", {})
        models = postprocessing.get("models", {})
        model_data = models.get(best_model, {})
        training = model_data.get("training", {})
    else:
        model_generation = item_data.get("model_generation", {})
        models = model_generation.get("models", {})
        model_data = models.get(best_model, {})
        training = model_data.get("training", {})
    
    total_days = training.get("total_days", 0)
    active_days = training.get("active_days", 0)
    
    # Fallback to compiled_results if training data not available
    if total_days == 0:
        compiled_results = item_data.get("compiled_results", {})
        if compiled_results and not compiled_results.get("error", False):
            original_summary = compiled_results.get("original_summary", {})
            total_days = original_summary.get("total_days", 0)
            active_days = original_summary.get("active_days", 0)
    
    return total_days, active_days


def generate_report(tracking_dir: Path, output_file: Path):
    """
    Generate Excel report from restaurant tracking files.
    
    Args:
        tracking_dir: Directory containing restaurant tracking JSON files
        output_file: Path to output Excel file
    """
    LOGGER.info("=" * 80)
    LOGGER.info("Generating Pipeline Report")
    LOGGER.info("=" * 80)
    
    # Load all tracking files
    all_tracking_data = load_all_tracking_files(tracking_dir)
    
    if not all_tracking_data:
        LOGGER.error("No tracking files found. Cannot generate report.")
        return
    
    successful_items = []
    failed_items = []
    
    # Process each tracking file
    for tracking_data in all_tracking_data:
        metadata = tracking_data.get("_metadata", {})
        foodcourt_id = metadata.get("foodcourt_id", "")
        foodcourt_name = metadata.get("foodcourt_name", foodcourt_id)
        restaurant_id = metadata.get("restaurant_id", "")
        restaurant_name = metadata.get("restaurant_name", restaurant_id)
        
        # Process each item
        for item_id, item_data in tracking_data.items():
            if item_id == "_metadata":
                continue
            
            item_name = item_data.get("item_name", item_id)
            
            # Check if item is successful
            is_successful, error_reason = is_item_successful(item_data)
            
            if is_successful:
                # Find best model
                best_model, postprocessing_used, metrics = find_best_model(item_data)
                
                if not best_model:
                    failed_items.append({
                        "Foodcourt_ID": foodcourt_id,
                        "Foodcourt_Name": foodcourt_name,
                        "Rest_ID": restaurant_id,
                        "Rest_Name": restaurant_name,
                        "Item_ID": item_id,
                        "Item_Name": item_name,
                        "Reason": "No best model found"
                    })
                    continue
                
                # Get file paths
                training_path_str, validation_path_str = get_file_paths(item_data, best_model, postprocessing_used)
                
                # Create hyperlinks
                training_hyperlink = "N/A"
                validation_hyperlink = "N/A"
                
                if training_path_str:
                    try:
                        training_path = Path(training_path_str)
                        if training_path.exists():
                            training_hyperlink = create_excel_hyperlink(
                                training_path,
                                f"{best_model}_Training"
                            )
                    except Exception as e:
                        LOGGER.debug(f"Failed to create training hyperlink: {e}")
                
                if validation_path_str:
                    try:
                        validation_path = Path(validation_path_str)
                        if validation_path.exists():
                            validation_hyperlink = create_excel_hyperlink(
                                validation_path,
                                f"{best_model}_Validation"
                            )
                    except Exception as e:
                        LOGGER.debug(f"Failed to create validation hyperlink: {e}")
                
                # Get total_days and active_days
                total_days, active_days = get_total_and_active_days(item_data, best_model, postprocessing_used)
                
                # Get accuracy reasons (only if < 75%)
                accuracy_reasons = get_accuracy_reasons(item_data, best_model, postprocessing_used)
                
                successful_items.append({
                    "Foodcourt_ID": foodcourt_id,
                    "Foodcourt_Name": foodcourt_name,
                    "Rest_ID": restaurant_id,
                    "Rest_Name": restaurant_name,
                    "Item_ID": item_id,
                    "Item_Name": item_name,
                    "Total_Days": total_days,
                    "Active_Days": active_days,
                    "Best_Model": best_model,
                    "PostProcessing_Used": "Yes" if postprocessing_used else "No",
                    "Abs_Avg_Dev": metrics.get("abs_avg_deviation", 0.0),
                    "Abs_Avg_Acc_Pct": metrics.get("avg_abs_accuracy_pct", 0.0),
                    "Training_File": training_hyperlink,
                    "Validation_File": validation_hyperlink,
                    "Reason": accuracy_reasons if accuracy_reasons else ""
                })
            else:
                failed_items.append({
                    "Foodcourt_ID": foodcourt_id,
                    "Foodcourt_Name": foodcourt_name,
                    "Rest_ID": restaurant_id,
                    "Rest_Name": restaurant_name,
                    "Item_ID": item_id,
                    "Item_Name": item_name,
                    "Reason": error_reason or "Unknown error"
                })
    
    LOGGER.info(f"Found {len(successful_items)} successful items and {len(failed_items)} failed items")
    
    # Create Excel workbook
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Sheet 1: Successful Compilation
    ws_success = wb.create_sheet("Successful Compilation")
    
    if successful_items:
        df_success = pd.DataFrame(successful_items)
        
        # Write headers
        headers = list(df_success.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws_success.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_idx, row_data in enumerate(df_success.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                header = headers[col_idx - 1]
                cell = ws_success.cell(row=row_idx, column=col_idx)
                
                # Handle hyperlink cells specially
                if header in ["Training_File", "Validation_File"]:
                    if isinstance(value, str) and value.startswith("=HYPERLINK"):
                        # Set as formula for Excel hyperlink
                        cell.value = value
                        cell.font = Font(color="0563C1", underline="single")
                    else:
                        cell.value = value or "N/A"
                else:
                    cell.value = value
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = max(
                len(str(header)),
                max((len(str(row[col_idx - 1])) for row in df_success.values), default=0)
            )
            ws_success.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
    else:
        # No successful items
        ws_success.cell(row=1, column=1, value="No successful items found")
    
    # Sheet 2: Remaining Data
    ws_failed = wb.create_sheet("Remaining Data")
    
    if failed_items:
        df_failed = pd.DataFrame(failed_items)
        
        # Write headers
        headers = list(df_failed.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws_failed.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_idx, row_data in enumerate(df_failed.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws_failed.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = max(
                len(str(header)),
                max((len(str(row[col_idx - 1])) for row in df_failed.values), default=0)
            )
            ws_failed.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
    else:
        # No failed items
        ws_failed.cell(row=1, column=1, value="No failed items found")
    
    # Save workbook
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    
    LOGGER.info(f"Report saved to: {output_file}")
    LOGGER.info(f"  - Successful items: {len(successful_items)}")
    LOGGER.info(f"  - Failed items: {len(failed_items)}")
    LOGGER.info("=" * 80)


def main():
    """Main function to generate report."""
    import sys
    
    # Get paths
    script_dir = Path(__file__).parent
    tracking_base_dir = script_dir / "output_data" / "restaurant_tracking"
    
    # Try to detect pipeline type from directory structure
    pipeline_type = "FRID_LEVEL"  # Default
    if tracking_base_dir.exists():
        subdirs = [d for d in tracking_base_dir.iterdir() if d.is_dir()]
        if subdirs:
            # Use the first subdirectory found (usually FRID_LEVEL)
            pipeline_type = subdirs[0].name
    
    tracking_dir = tracking_base_dir / pipeline_type
    
    # Output file in output_data/pipeline_report/{pipeline_type}/
    output_base_dir = script_dir / "output_data" / "pipeline_report" / pipeline_type
    output_file = output_base_dir / f"pipeline_report_{pipeline_type}.xlsx"
    
    LOGGER.info(f"Tracking directory: {tracking_dir}")
    LOGGER.info(f"Pipeline type: {pipeline_type}")
    LOGGER.info(f"Output file: {output_file}")
    
    if not tracking_dir.exists():
        LOGGER.error(f"Tracking directory not found: {tracking_dir}")
        LOGGER.info("Available directories:")
        if tracking_base_dir.exists():
            for subdir in tracking_base_dir.iterdir():
                if subdir.is_dir():
                    LOGGER.info(f"  - {subdir.name}")
        return
    
    generate_report(tracking_dir, output_file)


if __name__ == "__main__":
    main()

