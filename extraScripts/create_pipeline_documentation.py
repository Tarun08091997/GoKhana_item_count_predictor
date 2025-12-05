"""
Create comprehensive Excel documentation of the data pipeline.
Documents each component, what it does, and what data transformations occur.
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_FILE = Path("pipeline_documentation.xlsx")

# Define pipeline stages with detailed information
PIPELINE_STAGES = [
    {
        "Stage": "1. Fetch Restaurant Orders",
        "Component": "fetch_restaurant_orders.py",
        "Purpose": "Fetch raw order data from MongoDB",
        "Input": "MongoDB collection: entityrecord (foodorder data)",
        "Output": "Parquet/CSV files in input_data/fetched_data/{foodcourt_id}/{restaurant_id}.parquet",
        "Columns_Added": "placedtime, pickupdatetime, is_preorder, orderstatus, parentId (restaurant_id), entityId (foodcourt_id)",
        "Columns_Deleted": "None (raw data extraction)",
        "Columns_Modified": "None (raw data extraction)",
        "Transformations": "Extracts orders from MongoDB, separates preorders (pickupdatetime) from regular orders (placedtime), filters by orderstatus='completed'",
        "Data_Filtering": "Only completed orders are fetched",
        "File_Format": "Parquet (or CSV)",
    },
    {
        "Stage": "2. Enrich Restaurant Order Record",
        "Component": "enrich_restaurant_order_record.py",
        "Purpose": "Enrich order data with weather, holidays, and aggregate to daily item counts",
        "Input": "Parquet/CSV files from Stage 1",
        "Output": "CSV files in input_data/Model Training/enriched_data/{foodcourt_id}/{restaurant_id}.csv",
        "Columns_Added": "date (IST normalized, daily granularity), menuitemid, itemname, count (aggregated daily from orders), price (calculated from order totals), isVeg, isSpicy, foodcourtid, foodcourtname, restaurant, restaurantname, is_mon, is_tue, is_wed, is_thu, is_fri, is_sat, is_sun, is_holiday (derived from holiday data)",
        "Columns_Deleted": "Raw order-level fields (placedtime, pickupdatetime, orderstatus, etc.), date_IST (dropped after IST normalization), holiday_name, holiday_day, holiday_type (dropped in reorder_final_columns), weather columns (temperature, humidity, precipitation - removed in final output), is_preorder (aggregated away)",
        "Columns_Modified": "date: normalized to IST timezone, aggregated to daily level",
        "Transformations": "1. Align dates to IST timezone\n2. Match items to menu metadata\n3. Enrich with weather data from MySQL\n4. Enrich with holiday data from CSV\n5. Aggregate orders to daily item counts\n6. Add temporal features (weekday, weekend, holiday flags)",
        "Data_Filtering": "None - all orders processed",
        "File_Format": "CSV",
    },
    {
        "Stage": "3. Preprocess Data",
        "Component": "preprocess_data.py",
        "Purpose": "Clean, filter, and prepare data for model training",
        "Input": "CSV files from Stage 2 (enriched_data)",
        "Output": "CSV files in input_data/Model Training/preprocessed_data/{foodcourt_id}/{restaurant_id}.csv",
        "Columns_Added": "predict_model (1=XGBoost, 2/3=Moving Average), avg_3_day, avg_7_day, lag_7_day, lag_14_day, lag_21_day, lag_28_day, avg_1_month, avg_2_month, avg_3_month, is_mon, is_tue, is_wed, is_thu, is_fri, is_sat, is_sun",
        "Columns_Deleted": "holiday_name, holiday_day, holiday_type (kept is_holiday flag), date_IST (redundant after IST normalization), beverage/MRP items (rows filtered out completely and added to discard_report.csv)",
        "Columns_Modified": "count: converted to numeric, date: standardized format, menuitemid: converted to string",
        "Transformations": "1. Filter out beverage/MRP items\n2. Remove leading zero-count rows per item\n3. Compute item statistics (span_days, recent_total, recent_sale_days)\n4. Assign predict_model (1, 2, or 3) based on data span and activity\n5. Remove duplicate date-item combinations\n6. Add temporal features (day-of-week flags)\n7. Add rolling averages and lag features\n8. Calculate item statistics for model assignment",
        "Data_Filtering": "Only beverage/MRP items are discarded (recorded in discard_report.csv). All other items kept regardless of data quality.",
        "File_Format": "CSV",
    },
    {
        "Stage": "4. Model Training",
        "Component": "model_training.py",
        "Purpose": "Train forecasting models (XGBoost or Moving Average) for each item",
        "Input": "CSV files from Stage 3 (preprocessed_data)",
        "Output": "Model files and validation results in input_data/Model Training/XGBoost/{foodcourt_id}/{restaurant_id}/{item_slug}/",
        "Columns_Added": "predicted_count, error, error_pct (in validation results), model_name, train_rows, train_rmse, train_rmspe",
        "Columns_Deleted": "None (all input columns preserved in outputs)",
        "Columns_Modified": "None",
        "Transformations": "1. Split data into training (before MODEL_DATE) and validation (7 days before MODEL_DATE)\n2. For predict_model=1: Train XGBoost Regressor\n3. For predict_model=2/3: Train WeeklyMovingAverage (decay-based and weekday-aware methods)\n4. Generate predictions for validation period\n5. Calculate error metrics (RMSE, RMSPE, percentage errors)\n6. Save model files (.pkl) and validation results (.csv)",
        "Data_Filtering": "None - all items processed. Items with minimal data assigned to model 3 (Moving Average).",
        "File_Format": "CSV (results), PKL (models)",
    },
    {
        "Stage": "5. Postprocess Predictions",
        "Component": "postprocess_predictions.py",
        "Purpose": "Redistribute 7-day predictions based on day-of-week patterns from training data",
        "Input": "Validation predictions from Stage 4",
        "Output": "Postprocessed predictions in same item folder: {item_slug}_validation_postprocessed.csv",
        "Columns_Added": "predicted_count_redistributed, day_of_week_pct, error_postprocessing, error_pct_postprocessing",
        "Columns_Deleted": "None",
        "Columns_Modified": "predicted_count: redistributed based on day-of-week percentages",
        "Transformations": "1. Load validation predictions and training data\n2. Calculate day-of-week percentages from last 3 months of training data\n3. Calculate total 7-day prediction sum\n4. Redistribute predictions across 7 days based on percentages\n5. Recalculate errors with redistributed predictions",
        "Data_Filtering": "None",
        "File_Format": "CSV",
    },
    {
        "Stage": "6. Compile Errors",
        "Component": "compile_errors.py",
        "Purpose": "Compile error metrics from all models into comprehensive summary",
        "Input": "Validation results from all models (XGBoost, decay, weekday, postprocessing)",
        "Output": "error_compilation_all_models.csv and .xlsx in XGBoost output directory",
        "Columns_Added": "foodcourt_id, restaurant_id, item_id, item_name, xgboost_avg_error, xgboost_avg_error_cap, decay_avg_error, decay_avg_error_cap, weekday_avg_error, weekday_avg_error_cap, postprocessing_avg_error, postprocessing_avg_error_cap, folder_path, folder_link (Excel hyperlink)",
        "Columns_Deleted": "Individual day-level details (aggregated to per-item averages)",
        "Columns_Modified": "error_pct values: capped at 100% for average calculations only",
        "Transformations": "1. Load validation results from all model types\n2. Calculate average error percentages per item per model\n3. Calculate capped average errors (capping at 100%)\n4. Compile into one row per item\n5. Add folder paths and Excel hyperlinks for navigation",
        "Data_Filtering": "None - aggregates all available model results",
        "File_Format": "CSV, Excel (.xlsx)",
    },
]

# Additional details about data flow
DATA_FLOW_DETAILS = [
    {
        "Stage": "Stage 1 → Stage 2",
        "Transformation": "Order-level → Daily aggregated",
        "Granularity_Change": "Individual orders → Daily item counts",
        "Key_Operations": "Group by date + menuitemid, sum(count), enrich with weather/holidays",
    },
    {
        "Stage": "Stage 2 → Stage 3",
        "Transformation": "Enriched daily → Preprocessed daily",
        "Granularity_Change": "Daily level (maintained)",
        "Key_Operations": "Filter beverages, remove leading zeros, add temporal features, assign models",
    },
    {
        "Stage": "Stage 3 → Stage 4",
        "Transformation": "Preprocessed data → Trained models + predictions",
        "Granularity_Change": "Daily level → Model artifacts + validation predictions",
        "Key_Operations": "Split train/validation, train models, generate predictions",
    },
    {
        "Stage": "Stage 4 → Stage 5",
        "Transformation": "Raw predictions → Redistributed predictions",
        "Granularity_Change": "Daily predictions (maintained)",
        "Key_Operations": "Apply day-of-week percentage redistribution",
    },
    {
        "Stage": "Stage 5 → Stage 6",
        "Transformation": "Individual predictions → Aggregated error summaries",
        "Granularity_Change": "Daily level → Item-level averages",
        "Key_Operations": "Calculate averages, cap errors, compile across models",
    },
]

# Column transformations detail
COLUMN_TRANSFORMATIONS = [
    {
        "Stage": "Stage 2 (Enrichment)",
        "Column": "date",
        "Transformation": "Normalized to IST timezone, aggregated from order timestamps",
        "From": "placedtime (non-preorder) or pickupdatetime (preorder)",
        "To": "date (IST, daily granularity)",
    },
    {
        "Stage": "Stage 2 (Enrichment)",
        "Column": "count",
        "Transformation": "Aggregated from individual orders to daily totals per item",
        "From": "Individual order records",
        "To": "Daily count per menuitemid",
    },
    {
        "Stage": "Stage 3 (Preprocessing)",
        "Column": "predict_model",
        "Transformation": "Assigned based on data span and activity (1=XGBoost, 2/3=Moving Average)",
        "From": "N/A (new column)",
        "To": "Integer: 1, 2, or 3",
    },
    {
        "Stage": "Stage 3 (Preprocessing)",
        "Column": "avg_7_day, lag_7_day, etc.",
        "Transformation": "Calculated rolling averages and lag features",
        "From": "count column",
        "To": "Derived features for model training",
    },
    {
        "Stage": "Stage 4 (Model Training)",
        "Column": "predicted_count",
        "Transformation": "Generated by trained model",
        "From": "Model predictions on validation data",
        "To": "Predicted daily counts",
    },
    {
        "Stage": "Stage 5 (Postprocessing)",
        "Column": "predicted_count_redistributed",
        "Transformation": "Redistributed based on day-of-week percentages",
        "From": "predicted_count from Stage 4",
        "To": "Day-of-week adjusted predictions",
    },
]

def create_excel_documentation():
    """Create comprehensive Excel documentation of the pipeline."""
    
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # Sheet 1: Pipeline Overview
        df_overview = pd.DataFrame(PIPELINE_STAGES)
        df_overview.to_excel(writer, sheet_name='Pipeline Overview', index=False)
        
        # Sheet 2: Data Flow
        df_flow = pd.DataFrame(DATA_FLOW_DETAILS)
        df_flow.to_excel(writer, sheet_name='Data Flow', index=False)
        
        # Sheet 3: Column Transformations
        df_cols = pd.DataFrame(COLUMN_TRANSFORMATIONS)
        df_cols.to_excel(writer, sheet_name='Column Transformations', index=False)
        
        # Sheet 4: Detailed Stage-by-Stage Changes
        detailed_changes = []
        for stage in PIPELINE_STAGES:
            detailed_changes.append({
                "Stage": stage["Stage"],
                "Component": stage["Component"],
                "What_Added": stage["Columns_Added"],
                "What_Deleted": stage["Columns_Deleted"],
                "What_Modified": stage["Columns_Modified"],
                "Key_Transformations": stage["Transformations"],
                "Filtering_Rules": stage["Data_Filtering"],
            })
        df_detailed = pd.DataFrame(detailed_changes)
        df_detailed.to_excel(writer, sheet_name='Detailed Changes', index=False)
    
    # Format the Excel file
    format_excel_file(OUTPUT_FILE)
    print(f"Pipeline documentation created: {OUTPUT_FILE}")


def format_excel_file(file_path: Path):
    """Apply formatting to the Excel file."""
    wb = load_workbook(file_path)
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Format data rows and adjust column widths
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width (with some padding, max 80 characters)
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    wb.save(file_path)


if __name__ == "__main__":
    create_excel_documentation()
    print("\nPipeline documentation Excel file created successfully!")
    print(f"Location: {OUTPUT_FILE.resolve()}")

